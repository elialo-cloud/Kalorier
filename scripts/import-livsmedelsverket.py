#!/usr/bin/env python3
"""Import Livsmedelsverket's official Excel export into Cloudflare D1.

The script deliberately maps source values without recalculating or altering them.
It creates a SQL file that can be executed with `wrangler d1 execute kalorier-db`.

Usage:
  python scripts/import-livsmedelsverket.py Livsmedelsdatabasen.xlsx --out import.sql

The Excel export should be downloaded directly from Livsmedelsverket's
"Sök näringsinnehåll" service. The importer accepts Swedish/English header
variants and ignores columns it does not need.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installera först: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

SOURCE = "Livsmedelsverkets Livsmedelsdatabas version 2026-07-01"

ALIASES = {
    "name": ["livsmedel", "livsmedelsnamn", "food", "food name", "namn"],
    "source_id": ["livsmedelsnummer", "food number", "food id", "id"],
    "kcal": ["energi (kcal)", "energi kcal", "kcal", "energy kcal", "energy (kcal)"],
    "protein": ["protein (g)", "protein g", "protein"],
    "carbs": ["kolhydrater (g)", "kolhydrater g", "kolhydrater", "carbohydrate (g)", "carbohydrates (g)"],
    "fat": ["fett (g)", "fett g", "fett", "fat (g)"],
    "fiber": ["fiber (g)", "fiber g", "fiber"],
    "category": ["livsmedelsgrupp", "food group", "kategori", "category"],
}


def norm(value: object) -> str:
    s = "" if value is None else str(value)
    s = s.strip().lower().replace("å", "a").replace("ä", "a").replace("ö", "o")
    s = re.sub(r"\s+", " ", s)
    return s


def sql_string(value: object) -> str:
    if value is None or value == "":
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def number(value: object) -> str:
    if value is None or value == "":
        return "0"
    s = str(value).strip().replace(",", ".")
    try:
        return str(float(s))
    except ValueError:
        return "0"


def find_columns(headers: list[object]) -> dict[str, int]:
    normalized = {norm(h): i for i, h in enumerate(headers) if h is not None}
    found: dict[str, int] = {}
    for key, aliases in ALIASES.items():
        for alias in aliases:
            if norm(alias) in normalized:
                found[key] = normalized[norm(alias)]
                break
    return found


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--out", type=Path, default=Path("import-livsmedelsverket.sql"))
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    cols = find_columns(headers)

    if "name" not in cols or "kcal" not in cols:
        print("Kunde inte hitta kolumnerna för livsmedelsnamn och energi.", file=sys.stderr)
        print("Hittade rubriker:", ", ".join(str(h) for h in headers if h is not None), file=sys.stderr)
        return 2

    lines = [
        "-- Generated from " + SOURCE,
        "-- Source values are inserted as supplied; no nutrition values are recalculated.",
        "BEGIN TRANSACTION;",
        "DELETE FROM foods WHERE source = " + sql_string(SOURCE) + ";",
    ]
    count = 0

    for row in rows:
        if not row:
            continue
        name = row[cols["name"]] if cols["name"] < len(row) else None
        if not name or not str(name).strip():
            continue
        source_id = row[cols["source_id"]] if "source_id" in cols and cols["source_id"] < len(row) else None
        category = row[cols["category"]] if "category" in cols and cols["category"] < len(row) else None
        kcal = row[cols["kcal"]] if cols["kcal"] < len(row) else None
        protein = row[cols["protein"]] if "protein" in cols and cols["protein"] < len(row) else None
        carbs = row[cols["carbs"]] if "carbs" in cols and cols["carbs"] < len(row) else None
        fat = row[cols["fat"]] if "fat" in cols and cols["fat"] < len(row) else None
        fiber = row[cols["fiber"]] if "fiber" in cols and cols["fiber"] < len(row) else None
        normalized = norm(name)
        lines.append(
            "INSERT INTO foods (name,name_normalized,category,kcal_per_100g,protein_per_100g,"
            "carbs_per_100g,fat_per_100g,fiber_per_100g,source,source_id,verified) VALUES ("
            + ",".join([
                sql_string(str(name).strip()), sql_string(normalized), sql_string(category),
                number(kcal), number(protein), number(carbs), number(fat), number(fiber),
                sql_string(SOURCE), sql_string(source_id), "1"
            ]) + ");"
        )
        count += 1

    lines += ["COMMIT;", f"-- Imported rows: {count}"]
    args.out.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Skapade {args.out} med {count} livsmedel.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
