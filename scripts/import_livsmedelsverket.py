#!/usr/bin/env python3
"""Prepare Livsmedelsverket data for the Kalorier D1 schema.

The generated SQL assumes the core D1 migration has already created `foods`
and `food_portions`. The importer deliberately creates no competing schema.
"""
from pathlib import Path
import os
import re
import sqlite3
import subprocess
import tempfile
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "livsmedelsdatabasen.xlsx"
VERSION = os.getenv("LIVSMEDELSVERKET_VERSION", "2026-07-01")
SOURCE = "Livsmedelsverket Livsmedelsdatabas"


def norm(v):
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return None


def find_header(ws):
    aliases = {
        "name": ["livsmedelsnamn", "namn", "livsmedel"],
        "id": ["livsmedelsnummer", "livsmedelsnr", "nummer"],
        "kcal": ["energi (kcal)", "energi kcal", "kcal"],
        "protein": ["protein"],
        "carbs": ["kolhydrater", "kolhydrat"],
        "fat": ["fett"],
        "fiber": ["fiber", "fibrer"],
        "category": ["kategori", "livsmedelsgrupp", "grupp"],
        "edible_state": ["ätlig del", "ätbar del", "edible state"],
        "preparation": ["tillagning", "beredning", "preparation"],
        "barcode": ["streckkod", "ean", "barcode"],
    }
    for r in range(1, min(ws.max_row, 60) + 1):
        vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        found = {}
        for key, aa in aliases.items():
            for c, v in enumerate(vals, 1):
                if any(a == v or a in v for a in aa):
                    found[key] = c
                    break
        if "name" in found and ("kcal" in found or "id" in found):
            return r, found
    raise RuntimeError("Could not identify Livsmedelsverket header row/columns")


# Exact local representation of migrations/0001_foods.sql. Keeping this here
# makes the generated SQL testable without ever inventing a different schema.
LOCAL_SCHEMA = """
CREATE TABLE foods(
  id TEXT PRIMARY KEY,
  name TEXT NOT NULL,
  category TEXT,
  kcal_per_100g REAL NOT NULL,
  protein_per_100g REAL NOT NULL DEFAULT 0,
  carbs_per_100g REAL NOT NULL DEFAULT 0,
  fat_per_100g REAL NOT NULL DEFAULT 0,
  fiber_per_100g REAL,
  edible_state TEXT,
  preparation TEXT,
  source TEXT,
  source_id TEXT,
  barcode TEXT,
  verified INTEGER NOT NULL DEFAULT 0,
  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE food_portions(
  id TEXT PRIMARY KEY,
  food_id TEXT NOT NULL REFERENCES foods(id) ON DELETE CASCADE,
  name TEXT NOT NULL,
  grams REAL NOT NULL,
  UNIQUE(food_id, name)
);
CREATE TABLE import_metadata(key TEXT PRIMARY KEY, value TEXT NOT NULL);
"""


def sql_quote(value):
    return "NULL" if value is None else "'" + str(value).replace("'", "''") + "'"


def main():
    if not XLSX.exists():
        raise FileNotFoundError(XLSX)

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    header, cols = find_header(ws)

    fd, dbpath = tempfile.mkstemp(suffix=".sqlite")
    os.close(fd)
    con = sqlite3.connect(dbpath)
    cur = con.cursor()
    try:
        cur.executescript(LOCAL_SCHEMA)
        cur.execute("INSERT INTO import_metadata VALUES (?, ?)", ("source", SOURCE))
        cur.execute("INSERT INTO import_metadata VALUES (?, ?)", ("version", VERSION))

        n = 0
        for row in ws.iter_rows(min_row=header + 1, values_only=True):
            def get(key):
                c = cols.get(key)
                return row[c - 1] if c and c <= len(row) else None

            name = get("name")
            if not name or not str(name).strip():
                continue

            source_id = get("id")
            # Livsmedelsverket's source id is stable; use it as the D1 row id.
            food_id = f"livs:{source_id}" if source_id not in (None, "") else f"livs:row:{n + 1}"
            cur.execute(
                """INSERT OR REPLACE INTO foods
                (id,name,category,kcal_per_100g,protein_per_100g,carbs_per_100g,
                 fat_per_100g,fiber_per_100g,edible_state,preparation,source,
                 source_id,barcode,verified,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",
                (
                    food_id,
                    str(name).strip(),
                    get("category"),
                    num(get("kcal")) or 0,
                    num(get("protein")) or 0,
                    num(get("carbs")) or 0,
                    num(get("fat")) or 0,
                    num(get("fiber")),
                    get("edible_state"),
                    get("preparation"),
                    SOURCE,
                    str(source_id) if source_id is not None else None,
                    str(get("barcode")).strip() if get("barcode") not in (None, "") else None,
                    1,
                ),
            )
            n += 1

        con.commit()
        out = ROOT / "build" / "livsmedelsverket-import.sql"
        out.parent.mkdir(exist_ok=True)
        dump = subprocess.check_output(["sqlite3", dbpath, ".dump"], text=True)
        inserts = [
            line for line in dump.splitlines()
            if line.startswith("INSERT INTO foods") or line.startswith("INSERT INTO food_portions")
        ]
        metadata = [
            f"INSERT INTO import_metadata(key,value) VALUES ('source',{sql_quote(SOURCE)});",
            f"INSERT INTO import_metadata(key,value) VALUES ('version',{sql_quote(VERSION)});",
        ]
        lines = [
            "BEGIN TRANSACTION;",
            f"DELETE FROM food_portions WHERE food_id IN (SELECT id FROM foods WHERE source = {sql_quote(SOURCE)});",
            f"DELETE FROM foods WHERE source = {sql_quote(SOURCE)};",
            "DELETE FROM import_metadata WHERE key IN ('source','version');",
            *inserts,
            *metadata,
            "COMMIT;",
        ]
        out.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(f"Prepared {n} foods -> {out}")
    finally:
        con.close()
        os.unlink(dbpath)


if __name__ == "__main__":
    main()
